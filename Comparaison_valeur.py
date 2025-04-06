import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# 1. Charger les données
df = pd.read_excel("C:/Users/Bordeaux Montaigne/Desktop/Article/Tableau_actuel.xlsx")

# 2. Liste des variables microanatomiques (ajuster selon vos besoins)
variables_micro = ['RMeanT', 'RMaxT', 'C', 'TC', '%Trab']

# 3. Nettoyage des données
data_clean = df[['Context'] + variables_micro].dropna()

# 4. Création du fichier Excel
wb = Workbook()
ws = wb.active
ws.title = "Stats_Microanatomiques"

# Style d'en-tête
header_font = Font(bold=True)

# 5. Génération des statistiques pour chaque variable
for var in variables_micro:
    # Statistiques descriptives
    stats = data_clean.groupby('Context')[var].describe(percentiles=[.25, .5, .75])

    # Renommage des colonnes en français
    stats = stats.rename(columns={
        'count': 'Effectif',
        'mean': 'Moyenne',
        'std': 'Écart-type',
        'min': 'Minimum',
        '25%': '1er Quartile',
        '50%': 'Médiane',
        '75%': '3ème Quartile',
        'max': 'Maximum'
    })

    # Écriture dans Excel
    ws.append([f"STATISTIQUES POUR LA VARIABLE: {var}"])
    ws.append([])

    # Écriture des en-têtes
    headers = ['Contexte'] + list(stats.columns)
    ws.append(headers)

    # Mise en forme des en-têtes
    for cell in ws[ws.max_row]:
        cell.font = header_font

    # Écriture des données
    for context in stats.index:
        row = [context] + [round(val, 3) if isinstance(val, (float, int)) else val for val in stats.loc[context]]
        ws.append(row)

    # Ajout d'espace entre les variables
    ws.append([])
    ws.append([])

# 6. Ajustement automatique des colonnes
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

# 7. Sauvegarde
output_path = "Stats_Microanatomiques_par_Contexte.xlsx"
wb.save(output_path)
print(f"Rapport généré avec succès: {output_path}")